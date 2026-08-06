# ui/results_page.py
# ============================================================
# In-app results viewer for the log analyzer.
#
# Replaces "go open the spreadsheet" with a real, Excel-like
# grid inside the app:
#
#   * virtualised canvas grid (only visible rows are drawn, so
#     100k rows scroll as fast as 100)
#   * cell / range selection, arrow-key navigation, Ctrl+C copy
#     as TSV (pastes straight back into Excel)
#   * drag-to-resize columns, double-click divider to auto-fit
#   * per-column Excel-style filter dropdowns
#   * global search + issue severity filter + sortable headers
#   * frozen row-number column, frozen two-level header
#
# It reads the .xlsx that analyze_logs() already writes, so the
# analysis logic is untouched:
#   row 1  -> merged group headers (IXL Log, Wireshark Log, ...)
#   row 2  -> column headers
#   row 3+ -> data
#   fills FFFF00 / FF0000 / FFA500 -> issue highlights
#
# Usage inside the app (main.py already does this):
#   page = ResultsPage(container, controller=app)
#   ...later...
#   page.load_results(output_file_path)
#
# Standalone:
#   python -m ui.results_page  path\to\output.xlsx
# ============================================================

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from openpyxl import load_workbook

try:                       # normal in-app import
    from ui import palette as P
except ImportError:        # standalone / same-folder import
    import palette as P


ISSUE_LABELS = {
    "yellow": "Out of Order Sequence",
    "red":    "Missing Message",
    "orange": "Missing Packetswitch Data",
}

FILTER_OPTIONS = [
    "All rows",
    "Issues only",
    "Out of Order Sequence",
    "Missing Message",
    "Missing Packetswitch Data",
]

_LABEL_TO_KEY = {v: k for k, v in ISSUE_LABELS.items()}

BLANK_LABEL = "(blank)"


# ============================================================
# Workbook -> model
# ============================================================
def _fill_to_issue(cell):
    """Map an openpyxl cell fill to an issue key ('yellow'/'red'/'orange')."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type != "solid":
            return None
        rgb = fill.start_color.rgb
        if not isinstance(rgb, str):
            return None
        rgb = rgb[-6:].upper()
        if rgb == "FFFF00":
            return "yellow"
        if rgb == "FF0000":
            return "red"
        if rgb == "FFA500":
            return "orange"
    except Exception:
        pass
    return None


def empty_model():
    return {"columns": [], "groups": [], "rows": [], "highlights": {},
            "row_issues": [], "source": ""}


def load_results_from_workbook(xlsx_path, sheet_name=None):
    """
    Parse the analyzer's output workbook into a plain model dict:
      columns    : [str, ...]                      (row 2)
      groups     : [(label, start_idx, end_idx)]   (row 1, 0-based, inclusive)
      rows       : [[str, ...], ...]               (row 3+)
      highlights : {(row_idx, col_idx): issue_key}
      row_issues : [set(issue_key), ...]           (precomputed per row)
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    ncols = ws.max_column

    # --- group labels (row 1), expanded from merged ranges
    group_label = [""] * (ncols + 2)                    # 1-based
    for rng in ws.merged_cells.ranges:
        if rng.min_row == 1 and rng.max_row == 1:
            val = ws.cell(row=1, column=rng.min_col).value
            for c in range(rng.min_col, min(rng.max_col, ncols) + 1):
                group_label[c] = str(val) if val is not None else ""
    for c in range(1, ncols + 1):                       # non-merged row-1 cells
        v = ws.cell(row=1, column=c).value
        if v is not None and not group_label[c]:
            group_label[c] = str(v)

    groups = []
    c = 1
    while c <= ncols:
        label = group_label[c]
        start = c
        while c + 1 <= ncols and group_label[c + 1] == label and label != "":
            c += 1
        groups.append((label, start - 1, c - 1))        # 0-based, inclusive
        c += 1

    # --- column headers (row 2)
    columns = []
    for c in range(1, ncols + 1):
        v = ws.cell(row=2, column=c).value
        columns.append(str(v) if v is not None else "")

    # --- data rows + highlights
    rows, highlights, row_issues = [], {}, []
    r_out = 0
    for r in range(3, ws.max_row + 1):
        vals, any_val, issues = [], False, set()
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            v = "" if cell.value is None else str(cell.value)
            if v.strip():
                any_val = True
            vals.append(v)
            issue = _fill_to_issue(cell)
            if issue:
                highlights[(r_out, c - 1)] = issue
                issues.add(issue)
        if not any_val:
            # drop fully-empty rows, and the highlights we just recorded for them
            for cc in range(ncols):
                highlights.pop((r_out, cc), None)
            continue
        rows.append(vals)
        row_issues.append(issues)
        r_out += 1

    return {
        "columns": columns,
        "groups": groups,
        "rows": rows,
        "highlights": highlights,
        "row_issues": row_issues,
        "source": os.path.abspath(xlsx_path),
    }


# ============================================================
# Small themed widgets
# ============================================================
def _flat_button(parent, text, command, primary=False, width=None):
    bg = P.ACCENT if primary else P.SURFACE_ALT
    hov = P.ACCENT_HOVER if primary else P.BORDER
    fg = "#ffffff" if primary else P.FG
    b = tk.Button(parent, text=text, command=command, font=P.FONT,
                  relief="flat", bd=0, bg=bg, fg=fg,
                  activebackground=hov, activeforeground="#ffffff",
                  cursor="hand2", padx=12, pady=4,
                  highlightthickness=0)
    if width:
        b.configure(width=width)
    b.bind("<Enter>", lambda e: b.configure(bg=hov))
    b.bind("<Leave>", lambda e: b.configure(bg=bg))
    return b


class _FilterPopup(tk.Toplevel):
    """Excel-style value filter for one column."""

    def __init__(self, grid, ci, x_root, y_root):
        super().__init__(grid)
        self.grid_widget = grid
        self.ci = ci

        self.overrideredirect(True)
        self.configure(bg=P.BORDER)
        self.geometry(f"+{int(x_root)}+{int(y_root)}")

        wrap = tk.Frame(self, bg=P.SURFACE_ALT)
        wrap.pack(fill="both", expand=True, padx=1, pady=1)

        col_name = grid.model["columns"][ci] or f"Column {ci + 1}"
        tk.Label(wrap, text=col_name, font=P.FONT_BOLD, bg=P.SURFACE_ALT,
                 fg=P.FG_STRONG, anchor="w").pack(fill="x", padx=8, pady=(6, 2))

        sort_bar = tk.Frame(wrap, bg=P.SURFACE_ALT)
        sort_bar.pack(fill="x", padx=8, pady=(0, 4))
        _flat_button(sort_bar, "Sort A→Z",
                     lambda: self._sort(False)).pack(side="left", padx=(0, 4))
        _flat_button(sort_bar, "Sort Z→A",
                     lambda: self._sort(True)).pack(side="left")

        self._search = tk.StringVar()
        ent = tk.Entry(wrap, textvariable=self._search, font=P.FONT,
                       bg=P.BG, fg=P.FG, insertbackground=P.FG,
                       relief="flat", highlightthickness=1,
                       highlightbackground=P.BORDER, highlightcolor=P.ACCENT)
        ent.pack(fill="x", padx=8, pady=(0, 4), ipady=3)
        self._search.trace_add("write", lambda *_: self._refill())

        list_wrap = tk.Frame(wrap, bg=P.SURFACE_ALT)
        list_wrap.pack(fill="both", expand=True, padx=8)
        self.lb = tk.Listbox(list_wrap, selectmode="extended", height=12,
                             font=P.FONT, bg=P.BG, fg=P.FG,
                             selectbackground=P.ACCENT, selectforeground="#ffffff",
                             activestyle="none", relief="flat",
                             highlightthickness=1, highlightbackground=P.BORDER,
                             exportselection=False)
        sb = ttk.Scrollbar(list_wrap, orient="vertical", command=self.lb.yview)
        self.lb.configure(yscrollcommand=sb.set)
        self.lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        tk.Label(wrap, text="Ctrl / Shift click to multi-select",
                 font=P.FONT_SMALL, bg=P.SURFACE_ALT, fg=P.FG_MUTED,
                 anchor="w").pack(fill="x", padx=8, pady=(3, 0))

        btns = tk.Frame(wrap, bg=P.SURFACE_ALT)
        btns.pack(fill="x", padx=8, pady=6)
        _flat_button(btns, "All", self._select_all).pack(side="left", padx=(0, 4))
        _flat_button(btns, "None", self._select_none).pack(side="left")
        _flat_button(btns, "Apply", self._apply, primary=True).pack(side="right")
        _flat_button(btns, "Cancel", self.destroy).pack(side="right", padx=(0, 4))

        self._all_values = grid.distinct_values(ci)
        self._refill()

        self.transient(grid.winfo_toplevel())
        self.grab_set()
        ent.focus_set()
        self.bind("<Escape>", lambda e: self.destroy())
        self.bind("<Return>", lambda e: self._apply())
        self.bind("<FocusOut>", self._maybe_close)

    # -- helpers -------------------------------------------------
    def _maybe_close(self, _e=None):
        # closing on focus loss makes it feel like a real dropdown
        try:
            if self.focus_get() is None:
                self.destroy()
        except Exception:
            pass

    def _refill(self):
        needle = self._search.get().strip().lower()
        active = self.grid_widget.col_filters.get(self.ci)
        self._shown = [v for v in self._all_values
                       if not needle or needle in v.lower()]
        self.lb.delete(0, "end")
        for i, v in enumerate(self._shown):
            self.lb.insert("end", v if v else BLANK_LABEL)
            if active is None or v in active:
                self.lb.selection_set(i)

    def _select_all(self):
        self.lb.selection_set(0, "end")

    def _select_none(self):
        self.lb.selection_clear(0, "end")

    def _sort(self, desc):
        self.grid_widget.sort_by(self.ci, desc)
        self.destroy()

    def _apply(self):
        picked = {self._shown[i] for i in self.lb.curselection()}
        # values hidden by the popup search keep whatever they had before
        needle = self._search.get().strip().lower()
        if needle:
            prev = self.grid_widget.col_filters.get(self.ci)
            hidden = [v for v in self._all_values if needle not in v.lower()]
            for v in hidden:
                if prev is None or v in prev:
                    picked.add(v)
        if len(picked) == len(self._all_values):
            self.grid_widget.col_filters.pop(self.ci, None)     # no filter
        else:
            self.grid_widget.col_filters[self.ci] = picked
        self.grid_widget.apply_filters()
        self.destroy()


# ============================================================
# The grid
# ============================================================
class ExcelGrid(tk.Frame):
    """Virtualised, Excel-like table drawn on a canvas."""

    MIN_COL_W = 44
    MAX_COL_W = 420
    RESIZE_ZONE = 4
    FILTER_ZONE = 16

    def __init__(self, master, on_select=None, on_view_change=None,
                 frozen_cols=1, **kw):
        super().__init__(master, bg=P.BORDER, **kw)

        self.on_select = on_select
        self.on_view_change = on_view_change
        self.frozen_cols = frozen_cols

        self.model = empty_model()
        self.col_widths = []
        self.col_filters = {}          # {col_index: set(allowed raw values)}
        self._distinct_cache = {}

        self._search_text = ""
        self._severity = FILTER_OPTIONS[0]
        self._sort_col = None
        self._sort_desc = False

        self._view = []                # view row -> model row index
        self._xoff = 0
        self._yrow = 0.0

        self._anchor = None            # (view_row, col)
        self._focus = None             # (view_row, col)

        self._drag_col = None          # column being resized
        self._drag_x0 = 0
        self._drag_w0 = 0
        self._press = None             # ("header", ci, y) pending click

        # ---- widgets ----
        self.hdr = tk.Canvas(self, height=P.GROUP_H + P.COL_H, bg=P.HEADER_BG,
                             highlightthickness=0)
        self.body = tk.Canvas(self, bg=P.SURFACE, highlightthickness=0,
                              takefocus=True)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self._on_vscroll)
        self.hsb = ttk.Scrollbar(self, orient="horizontal", command=self._on_hscroll)
        corner = tk.Frame(self, bg=P.HEADER_BG)

        self.hdr.grid(row=0, column=0, sticky="ew")
        corner.grid(row=0, column=1, sticky="nsew")
        self.body.grid(row=1, column=0, sticky="nsew")
        self.vsb.grid(row=1, column=1, sticky="ns")
        self.hsb.grid(row=2, column=0, sticky="ew")
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)

        # ---- bindings ----
        self.body.bind("<Configure>", lambda e: self.redraw())
        self.body.bind("<MouseWheel>", self._on_wheel)
        self.body.bind("<Shift-MouseWheel>", self._on_shift_wheel)
        self.body.bind("<Button-4>", lambda e: self._scroll_rows(-3))
        self.body.bind("<Button-5>", lambda e: self._scroll_rows(3))
        self.body.bind("<Button-1>", self._on_body_press)
        self.body.bind("<B1-Motion>", self._on_body_drag)
        self.body.bind("<Shift-Button-1>", self._on_body_shift_click)
        self.body.bind("<Key>", self._on_key)
        self.body.bind("<Control-c>", lambda e: (self.copy_selection(), "break")[1])
        self.body.bind("<Control-C>", lambda e: (self.copy_selection(), "break")[1])
        self.body.bind("<Control-a>", lambda e: (self.select_all(), "break")[1])

        self.hdr.bind("<Motion>", self._on_hdr_motion)
        self.hdr.bind("<Button-1>", self._on_hdr_press)
        self.hdr.bind("<B1-Motion>", self._on_hdr_drag)
        self.hdr.bind("<ButtonRelease-1>", self._on_hdr_release)
        self.hdr.bind("<Double-Button-1>", self._on_hdr_double)
        self.hdr.bind("<MouseWheel>", self._on_shift_wheel)

    # ------------------------------------------------------------------
    # model
    # ------------------------------------------------------------------
    def set_model(self, model):
        self.model = model or empty_model()
        self.col_filters = {}
        self._distinct_cache = {}
        self._sort_col = None
        self._sort_desc = False
        self._xoff = 0
        self._yrow = 0.0
        self._anchor = self._focus = None
        self.col_widths = self._auto_widths()
        self.apply_filters()

    def _auto_widths(self):
        px = 7                                    # ~px per char at 9pt
        widths = []
        rows = self.model["rows"]
        sample = rows[:500]
        for ci, col in enumerate(self.model["columns"]):
            w = len(col) + 3                      # room for the ▾ glyph
            for row in sample:
                v = row[ci] if ci < len(row) else ""
                if v:
                    w = max(w, min(len(v.splitlines()[0]), 40))
            widths.append(max(self.MIN_COL_W, min(self.MAX_COL_W, w * px + 20)))
        return widths

    def distinct_values(self, ci):
        if ci in self._distinct_cache:
            return self._distinct_cache[ci]
        seen = set()
        for row in self.model["rows"]:
            seen.add(row[ci] if ci < len(row) else "")
        vals = sorted(seen, key=lambda s: (s == "", s.lower()))
        self._distinct_cache[ci] = vals
        return vals

    # ------------------------------------------------------------------
    # filtering / sorting
    # ------------------------------------------------------------------
    def set_search(self, text):
        self._search_text = (text or "").strip().lower()
        self.apply_filters()

    def set_severity(self, option):
        self._severity = option
        self.apply_filters()

    def clear_all_filters(self):
        self.col_filters = {}
        self._search_text = ""
        self._severity = FILTER_OPTIONS[0]
        self._sort_col = None
        self._sort_desc = False
        self.apply_filters()

    def sort_by(self, ci, desc):
        self._sort_col, self._sort_desc = ci, desc
        self.apply_filters()

    def apply_filters(self):
        rows = self.model["rows"]
        issues_per_row = self.model.get("row_issues") or [set()] * len(rows)
        sev = self._severity
        needle = self._search_text
        colf = self.col_filters

        wanted = _LABEL_TO_KEY.get(sev)

        idxs = []
        for i in range(len(rows)):
            row = rows[i]
            if sev != FILTER_OPTIONS[0]:
                ri = issues_per_row[i]
                if sev == "Issues only":
                    if not ri:
                        continue
                elif wanted not in ri:
                    continue
            if colf:
                bad = False
                for ci, allowed in colf.items():
                    v = row[ci] if ci < len(row) else ""
                    if v not in allowed:
                        bad = True
                        break
                if bad:
                    continue
            if needle and not any(needle in v.lower() for v in row if v):
                continue
            idxs.append(i)

        if self._sort_col is not None and self._sort_col < len(self.model["columns"]):
            col = self._sort_col

            def key(i):
                v = rows[i][col] if col < len(rows[i]) else ""
                try:
                    return (0, float(v), "")
                except (TypeError, ValueError):
                    return (1, 0.0, v.lower())

            idxs.sort(key=key, reverse=self._sort_desc)

        self._view = idxs
        self._yrow = 0.0
        self._anchor = self._focus = None
        if self.on_view_change:
            self.on_view_change(len(idxs), len(rows))
        if self.on_select:
            self.on_select(None)
        self.redraw()

    # ------------------------------------------------------------------
    # geometry helpers
    # ------------------------------------------------------------------
    def _ncols(self):
        return len(self.col_widths)

    def _frozen_w(self):
        return sum(self.col_widths[:self.frozen_cols])

    def _total_w(self):
        return sum(self.col_widths)

    def _col_x(self, ci):
        return sum(self.col_widths[:ci])

    def _draw_x(self, ci):
        """Canvas x of column ci, accounting for the frozen pane."""
        if ci < self.frozen_cols:
            return self._col_x(ci)
        return self._col_x(ci) - self._xoff

    def _visible_rows(self):
        h = max(self.body.winfo_height(), 1)
        return max(int(h // P.ROW_H), 1)

    def _max_first_row(self):
        return max(len(self._view) - self._visible_rows(), 0)

    def _max_xoff(self):
        w = max(self.body.winfo_width(), 1)
        return max(self._total_w() - w, 0)

    def _col_at_x(self, x):
        fw = self._frozen_w()
        if x < fw:
            acc = 0
            for ci in range(self.frozen_cols):
                if acc <= x < acc + self.col_widths[ci]:
                    return ci
                acc += self.col_widths[ci]
            return max(self.frozen_cols - 1, 0)
        xx = x + self._xoff
        acc = 0
        for ci in range(self._ncols()):
            if acc <= xx < acc + self.col_widths[ci]:
                return ci if ci >= self.frozen_cols else None
            acc += self.col_widths[ci]
        return None

    def _boundary_at_x(self, x):
        """Return column index whose RIGHT edge is within the resize zone of x."""
        fw = self._frozen_w()
        acc = 0
        for ci in range(self._ncols()):
            acc += self.col_widths[ci]
            edge = acc if ci < self.frozen_cols else acc - self._xoff
            if ci >= self.frozen_cols and edge < fw:
                continue
            if abs(x - edge) <= self.RESIZE_ZONE:
                return ci
        return None

    # ------------------------------------------------------------------
    # scrolling
    # ------------------------------------------------------------------
    def _on_vscroll(self, *args):
        if args[0] == "moveto":
            self._yrow = float(args[1]) * self._max_first_row()
        elif args[0] == "scroll":
            step = int(args[1])
            if args[2] == "pages":
                step *= self._visible_rows()
            self._yrow += step
        self._yrow = min(max(self._yrow, 0), self._max_first_row())
        self.redraw()

    def _on_hscroll(self, *args):
        if args[0] == "moveto":
            self._xoff = int(float(args[1]) * max(self._total_w(), 1))
        elif args[0] == "scroll":
            w = max(self.body.winfo_width(), 1)
            self._xoff += int(args[1]) * (w if args[2] == "pages" else 48)
        self._xoff = min(max(self._xoff, 0), self._max_xoff())
        self.redraw()

    def _scroll_rows(self, n):
        self._yrow = min(max(self._yrow + n, 0), self._max_first_row())
        self.redraw()

    def _on_wheel(self, event):
        self._scroll_rows(-3 if event.delta > 0 else 3)
        return "break"

    def _on_shift_wheel(self, event):
        self._xoff = min(max(self._xoff + (-60 if event.delta > 0 else 60), 0),
                         self._max_xoff())
        self.redraw()
        return "break"

    def _update_scrollbars(self):
        n = max(len(self._view), 1)
        vis = self._visible_rows()
        self.vsb.set(self._yrow / n, min((self._yrow + vis) / n, 1.0))

        w = max(self.body.winfo_width(), 1)
        tw = max(self._total_w(), 1)
        if tw <= w:
            self.hsb.set(0, 1)
        else:
            self.hsb.set(self._xoff / tw, min((self._xoff + w) / tw, 1.0))

    def _ensure_visible(self, vr, ci):
        first = int(self._yrow)
        vis = self._visible_rows()
        if vr < first:
            self._yrow = vr
        elif vr >= first + vis:
            self._yrow = vr - vis + 1
        self._yrow = min(max(self._yrow, 0), self._max_first_row())

        if ci is not None and ci >= self.frozen_cols:
            fw = self._frozen_w()
            w = max(self.body.winfo_width(), 1)
            left = self._col_x(ci)
            right = left + self.col_widths[ci]
            if left - self._xoff < fw:
                self._xoff = left - fw
            elif right - self._xoff > w:
                self._xoff = right - w
            self._xoff = min(max(self._xoff, 0), self._max_xoff())

    # ------------------------------------------------------------------
    # drawing
    # ------------------------------------------------------------------
    def redraw(self):
        self._draw_header()
        self._draw_body()
        self._update_scrollbars()

    def _shorten(self, text, width):
        max_chars = max(int((width - 12) / 6.6), 1)
        if len(text) > max_chars:
            return text[:max_chars - 1] + "…"
        return text

    def _draw_header(self):
        c = self.hdr
        c.delete("all")
        gh, ch = P.GROUP_H, P.COL_H
        fw = self._frozen_w()
        w_vis = max(c.winfo_width(), 1)

        def draw_group(label, s, e, frozen):
            x0 = self._draw_x(s)
            x1 = x0 + sum(self.col_widths[s:e + 1])
            if x1 < 0 or x0 > w_vis:
                return
            c.create_rectangle(x0, 0, x1, gh, fill=P.GROUP_BG,
                               outline=P.BORDER, width=1)
            if label:
                c.create_text((x0 + x1) / 2, gh / 2, text=label,
                              font=P.FONT_BOLD, fill=P.GROUP_FG,
                              width=max(x1 - x0 - 8, 10))

        def draw_col(ci, frozen):
            x = self._draw_x(ci)
            w = self.col_widths[ci]
            if x + w < 0 or x > w_vis:
                return
            c.create_rectangle(x, gh, x + w, gh + ch, fill=P.HEADER_BG,
                               outline=P.BORDER, width=1)
            label = self.model["columns"][ci]
            if ci == self._sort_col:
                label += "  ▼" if self._sort_desc else "  ▲"
            c.create_text(x + 6, gh + ch / 2, text=self._shorten(label, w - 14),
                          anchor="w", font=P.FONT_BOLD, fill=P.HEADER_FG)
            filtered = ci in self.col_filters
            c.create_text(x + w - 8, gh + ch / 2, text="▼" if filtered else "▾",
                          anchor="e", font=P.FONT_SMALL if filtered else P.FONT,
                          fill=P.ACCENT_HOVER if filtered else P.FG_MUTED)

        # scrollable pane
        for label, s, e in self.model["groups"]:
            if e >= self.frozen_cols:
                draw_group(label, max(s, self.frozen_cols), e, False)
        for ci in range(self.frozen_cols, self._ncols()):
            draw_col(ci, False)

        # frozen pane on top
        if fw:
            c.create_rectangle(0, 0, fw, gh + ch, fill=P.HEADER_BG, outline="")
            for label, s, e in self.model["groups"]:
                if s < self.frozen_cols:
                    draw_group(label, s, min(e, self.frozen_cols - 1), True)
            for ci in range(self.frozen_cols):
                draw_col(ci, True)
            c.create_line(fw, 0, fw, gh + ch, fill=P.ACCENT_DIM, width=1)

    def _draw_body(self):
        c = self.body
        c.delete("all")
        rows = self.model["rows"]
        hl = self.model["highlights"]
        rh = P.ROW_H
        fw = self._frozen_w()
        w_vis = max(c.winfo_width(), 1)
        h_vis = max(c.winfo_height(), 1)

        first = int(self._yrow)
        last = min(first + self._visible_rows() + 1, len(self._view))
        sel = self._sel_rect()

        def base_bg(vr):
            if sel and sel[0] <= vr <= sel[1]:
                return P.SELECT_BG
            return P.ROW_ILL if (vr % 2) else P.ROW_EVEN

        def cell_bg(vr, ci, mi):
            issue = hl.get((mi, ci))
            if issue:
                return P.ISSUE_BG[issue], P.ISSUE_FG[issue]
            if sel and sel[0] <= vr <= sel[1] and sel[2] <= ci <= sel[3]:
                return P.SELECT_BG, P.FG_STRONG
            return None, P.FG

        def draw_cell(vr, mi, ci, y):
            x = self._draw_x(ci)
            w = self.col_widths[ci]
            if x + w < 0 or x > w_vis:
                return
            bg, fg = cell_bg(vr, ci, mi)
            if bg:
                c.create_rectangle(x + 1, y + 1, x + w - 1, y + rh - 1,
                                   fill=bg, outline="")
            row = rows[mi]
            v = row[ci] if ci < len(row) else ""
            if v:
                text = v.splitlines()[0]
                c.create_text(x + 6, y + rh / 2, anchor="w",
                              text=self._shorten(text, w),
                              font=P.FONT, fill=fg)
            c.create_line(x + w, y, x + w, y + rh, fill=P.GRID_LINE)

        # ---- scrollable pane
        y = 0
        for vr in range(first, last):
            mi = self._view[vr]
            c.create_rectangle(0, y, w_vis, y + rh, fill=base_bg(vr), outline="")
            for ci in range(self.frozen_cols, self._ncols()):
                draw_cell(vr, mi, ci, y)
            c.create_line(0, y + rh, w_vis, y + rh, fill=P.GRID_LINE)
            y += rh

        # ---- frozen pane painted over the top
        if fw:
            y = 0
            for vr in range(first, last):
                mi = self._view[vr]
                c.create_rectangle(0, y, fw, y + rh, fill=base_bg(vr), outline="")
                for ci in range(self.frozen_cols):
                    draw_cell(vr, mi, ci, y)
                c.create_line(0, y + rh, fw, y + rh, fill=P.GRID_LINE)
                y += rh
            c.create_line(fw, 0, fw, h_vis, fill=P.ACCENT_DIM, width=1)

        # ---- focus box
        if self._focus:
            fr, fc = self._focus
            if first <= fr < last and fc is not None:
                fx = self._draw_x(fc)
                fy = (fr - first) * rh
                if fc >= self.frozen_cols and fx < fw:
                    pass                                    # hidden behind pane
                else:
                    c.create_rectangle(fx + 1, fy + 1,
                                       fx + self.col_widths[fc] - 1, fy + rh - 1,
                                       outline=P.FOCUS_BD, width=2)

        if not self._view:
            msg = ("No results loaded — run an analysis."
                   if not rows else "No rows match the current filters.")
            c.create_text(w_vis / 2, h_vis / 2, text=msg,
                          font=P.FONT, fill=P.FG_MUTED)

    # ------------------------------------------------------------------
    # selection
    # ------------------------------------------------------------------
    def _sel_rect(self):
        if not self._anchor or not self._focus:
            return None
        (r0, c0), (r1, c1) = self._anchor, self._focus
        if c0 is None or c1 is None:
            return None
        return (min(r0, r1), max(r0, r1), min(c0, c1), max(c0, c1))

    def _cell_at(self, event):
        vr = int(self._yrow) + int(event.y // P.ROW_H)
        if vr < 0 or vr >= len(self._view):
            return None
        ci = self._col_at_x(event.x)
        if ci is None:
            return None
        return (vr, ci)

    def _on_body_press(self, event):
        self.body.focus_set()
        cell = self._cell_at(event)
        if not cell:
            return
        self._anchor = self._focus = cell
        self._notify_select()
        self.redraw()

    def _on_body_shift_click(self, event):
        cell = self._cell_at(event)
        if not cell or not self._anchor:
            return self._on_body_press(event)
        self._focus = cell
        self._notify_select()
        self.redraw()
        return "break"

    def _on_body_drag(self, event):
        if not self._anchor:
            return
        vr = int(self._yrow) + int(event.y // P.ROW_H)
        vr = min(max(vr, 0), max(len(self._view) - 1, 0))
        ci = self._col_at_x(min(max(event.x, 0), max(self.body.winfo_width() - 1, 0)))
        if ci is None:
            ci = self._focus[1] if self._focus else 0
        self._focus = (vr, ci)
        self._ensure_visible(vr, ci)
        self._notify_select()
        self.redraw()

    def _notify_select(self):
        if self.on_select and self._focus and self._view:
            vr = self._focus[0]
            if 0 <= vr < len(self._view):
                self.on_select(self._view[vr])

    def select_all(self):
        if not self._view:
            return
        self._anchor = (0, 0)
        self._focus = (len(self._view) - 1, self._ncols() - 1)
        self.redraw()

    def _on_key(self, event):
        if not self._view:
            return
        k = event.keysym
        shift = bool(event.state & 0x0001)
        if self._focus is None:
            self._anchor = self._focus = (int(self._yrow), self.frozen_cols)

        vr, ci = self._focus
        moved = True
        if k == "Up":
            vr -= 1
        elif k == "Down":
            vr += 1
        elif k == "Left":
            ci -= 1
        elif k == "Right":
            ci += 1
        elif k == "Prior":
            vr -= self._visible_rows()
        elif k == "Next":
            vr += self._visible_rows()
        elif k == "Home":
            ci = 0
        elif k == "End":
            ci = self._ncols() - 1
        elif k == "Escape":
            self._anchor = self._focus = None
            self.redraw()
            return "break"
        else:
            moved = False

        if not moved:
            return

        vr = min(max(vr, 0), len(self._view) - 1)
        ci = min(max(ci, 0), self._ncols() - 1)
        self._focus = (vr, ci)
        if not shift:
            self._anchor = self._focus
        self._ensure_visible(vr, ci)
        self._notify_select()
        self.redraw()
        return "break"

    def copy_selection(self):
        rect = self._sel_rect()
        rows = self.model["rows"]
        if not rect:
            return
        r0, r1, c0, c1 = rect
        lines = []
        for vr in range(r0, r1 + 1):
            row = rows[self._view[vr]]
            cells = []
            for ci in range(c0, c1 + 1):
                v = row[ci] if ci < len(row) else ""
                cells.append(v.replace("\r\n", " ").replace("\n", " ").replace("\t", " "))
            lines.append("\t".join(cells))
        text = "\n".join(lines)
        self.clipboard_clear()
        self.clipboard_append(text)

    # ------------------------------------------------------------------
    # header interaction: resize / sort / filter
    # ------------------------------------------------------------------
    def _on_hdr_motion(self, event):
        if event.y < P.GROUP_H:
            self.hdr.configure(cursor="")
            return
        self.hdr.configure(
            cursor="sb_h_double_arrow" if self._boundary_at_x(event.x) is not None
            else "hand2")

    def _on_hdr_press(self, event):
        self._press = None
        if event.y < P.GROUP_H:
            return
        ci = self._boundary_at_x(event.x)
        if ci is not None:
            self._drag_col = ci
            self._drag_x0 = event.x
            self._drag_w0 = self.col_widths[ci]
            return
        ci = self._col_at_x(event.x)
        if ci is not None:
            self._press = (ci, event.x)

    def _on_hdr_drag(self, event):
        if self._drag_col is None:
            return
        ci = self._drag_col
        neww = self._drag_w0 + (event.x - self._drag_x0)
        self.col_widths[ci] = int(min(max(neww, self.MIN_COL_W), self.MAX_COL_W))
        self._xoff = min(self._xoff, self._max_xoff())
        self.redraw()

    def _on_hdr_release(self, event):
        if self._drag_col is not None:
            self._drag_col = None
            self._on_hdr_motion(event)
            return
        if not self._press:
            return
        ci, x0 = self._press
        self._press = None
        if abs(event.x - x0) > 3:
            return
        x = self._draw_x(ci)
        w = self.col_widths[ci]
        if event.x >= x + w - self.FILTER_ZONE:
            self._open_filter(ci)
        else:
            if self._sort_col == ci:
                if self._sort_desc:
                    self._sort_col, self._sort_desc = None, False
                else:
                    self._sort_desc = True
            else:
                self._sort_col, self._sort_desc = ci, False
            self.apply_filters()

    def _on_hdr_double(self, event):
        if event.y < P.GROUP_H:
            return
        ci = self._boundary_at_x(event.x)
        if ci is None:
            return "break"
        self.autofit(ci)
        return "break"

    def autofit(self, ci):
        px = 7
        w = len(self.model["columns"][ci]) + 3
        for mi in self._view[:2000]:
            row = self.model["rows"][mi]
            v = row[ci] if ci < len(row) else ""
            if v:
                w = max(w, min(len(v.splitlines()[0]), 40))
        self.col_widths[ci] = max(self.MIN_COL_W, min(self.MAX_COL_W, w * px + 20))
        self.redraw()

    def _open_filter(self, ci):
        x = self.hdr.winfo_rootx() + max(self._draw_x(ci), 0)
        y = self.hdr.winfo_rooty() + P.GROUP_H + P.COL_H
        _FilterPopup(self, ci, x, y)

    # ------------------------------------------------------------------
    # export
    # ------------------------------------------------------------------
    def export_view_csv(self, path):
        import csv
        rows = self.model["rows"]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(self.model["columns"])
            for mi in self._view:
                w.writerow(rows[mi])


# ============================================================
# The page
# ============================================================
class ResultsPage(tk.Frame):
    def __init__(self, master, controller=None, xlsx_path=None, model=None, **kw):
        super().__init__(master, bg=P.BG, **kw)
        self.controller = controller
        self.model = empty_model()

        self._build_ui()

        if model is not None:
            self.set_model(model)
        elif xlsx_path:
            self.load_results(xlsx_path)
        else:
            self.set_model(empty_model())

    # ---------------- public API ----------------
    def load_results(self, xlsx_path, sheet_name=None):
        """Load (or reload) the analyzer output workbook into the page."""
        try:
            model = load_results_from_workbook(xlsx_path, sheet_name=sheet_name)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read results file:\n{e}")
            return False
        self.set_model(model)
        return True

    def set_model(self, model):
        self.model = model or empty_model()
        self._src_lbl.config(
            text=os.path.basename(self.model.get("source", "")) or "no file loaded")
        self._search_var.set("")
        self._sev_var.set(FILTER_OPTIONS[0])
        self.grid_view.set_model(self.model)
        self._update_kpis()
        self._show_detail(None)

    def on_show(self):
        pass

    # ---------------- UI ----------------
    def _build_ui(self):
        # ---- title bar
        top = tk.Frame(self, bg=P.BG)
        top.pack(fill="x", padx=16, pady=(14, 8))
        tk.Label(top, text="Analysis Results", font=P.FONT_TITLE,
                 bg=P.BG, fg=P.FG_STRONG).pack(side="left")
        self._src_lbl = tk.Label(top, text="", font=P.FONT_SMALL,
                                 bg=P.BG, fg=P.FG_MUTED)
        self._src_lbl.pack(side="left", padx=(10, 0), pady=(7, 0))

        _flat_button(top, "Open Excel file", self._open_excel,
                     primary=True).pack(side="right")
        _flat_button(top, "Export view to CSV",
                     self._export_csv).pack(side="right", padx=(0, 8))
        _flat_button(top, "Copy selection",
                     lambda: self.grid_view.copy_selection()).pack(side="right",
                                                                   padx=(0, 8))

        # ---- KPI cards
        self._kpi_vals = {}
        kpi = tk.Frame(self, bg=P.BG)
        kpi.pack(fill="x", padx=16, pady=(0, 10))
        cards = [
            ("Total Rows", None, P.FG_STRONG, P.ACCENT, FILTER_OPTIONS[0]),
            (ISSUE_LABELS["yellow"], "yellow", P.ISSUE_FG["yellow"],
             P.ISSUE_DOT["yellow"], ISSUE_LABELS["yellow"]),
            (ISSUE_LABELS["red"], "red", P.ISSUE_FG["red"],
             P.ISSUE_DOT["red"], ISSUE_LABELS["red"]),
            (ISSUE_LABELS["orange"], "orange", P.ISSUE_FG["orange"],
             P.ISSUE_DOT["orange"], ISSUE_LABELS["orange"]),
        ]
        for label, key, fg, bar, filt in cards:
            card = tk.Frame(kpi, bg=P.SURFACE, highlightthickness=1,
                            highlightbackground=P.BORDER, cursor="hand2")
            card.pack(side="left", padx=(0, 10))
            strip = tk.Frame(card, bg=bar, width=4)
            strip.pack(side="left", fill="y")
            inner = tk.Frame(card, bg=P.SURFACE)
            inner.pack(side="left", padx=(12, 18), pady=6)
            val = tk.Label(inner, text="0", font=P.FONT_KPI, bg=P.SURFACE, fg=fg)
            val.pack(anchor="w")
            cap = tk.Label(inner, text=label, font=P.FONT_SMALL, bg=P.SURFACE,
                           fg=P.FG_MUTED)
            cap.pack(anchor="w")
            self._kpi_vals[key] = val
            for w in (card, inner, val, cap, strip):
                w.bind("<Button-1>", lambda e, f=filt: self._set_severity(f))
                w.bind("<Enter>", lambda e, c=card: c.configure(
                    highlightbackground=P.ACCENT))
                w.bind("<Leave>", lambda e, c=card: c.configure(
                    highlightbackground=P.BORDER))

        # ---- filter bar
        bar = tk.Frame(self, bg=P.BG)
        bar.pack(fill="x", padx=16, pady=(0, 8))

        tk.Label(bar, text="Search", font=P.FONT, bg=P.BG,
                 fg=P.FG).pack(side="left")
        self._search_var = tk.StringVar()
        ent = tk.Entry(bar, textvariable=self._search_var, font=P.FONT, width=34,
                       bg=P.SURFACE_ALT, fg=P.FG, insertbackground=P.FG,
                       relief="flat", highlightthickness=1,
                       highlightbackground=P.BORDER, highlightcolor=P.ACCENT)
        ent.pack(side="left", padx=(8, 18), ipady=3)
        self._search_var.trace_add(
            "write", lambda *_: self.grid_view.set_search(self._search_var.get()))

        tk.Label(bar, text="Show", font=P.FONT, bg=P.BG, fg=P.FG).pack(side="left")
        self._sev_var = tk.StringVar(value=FILTER_OPTIONS[0])
        combo = ttk.Combobox(bar, textvariable=self._sev_var, values=FILTER_OPTIONS,
                             state="readonly", width=26, font=P.FONT,
                             style="Dark.TCombobox")
        combo.pack(side="left", padx=(8, 18))
        combo.bind("<<ComboboxSelected>>",
                   lambda e: self._set_severity(self._sev_var.get()))

        _flat_button(bar, "Clear filters", self._clear_filters).pack(side="left")

        self._count_lbl = tk.Label(bar, text="", font=P.FONT, bg=P.BG,
                                   fg=P.FG_MUTED)
        self._count_lbl.pack(side="right")

        # ---- grid
        self.grid_view = ExcelGrid(self, on_select=self._show_detail,
                                   on_view_change=self._on_view_change,
                                   frozen_cols=1)
        self.grid_view.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # ---- legend + row details
        bottom = tk.Frame(self, bg=P.BG)
        bottom.pack(fill="x", padx=16, pady=(0, 12))

        legend = tk.Frame(bottom, bg=P.BG)
        legend.pack(side="left", anchor="n", pady=4)
        for key in ("yellow", "red", "orange"):
            sw = tk.Frame(legend, bg=P.ISSUE_DOT[key], width=12, height=12)
            sw.pack(side="left", padx=(0, 5))
            sw.pack_propagate(False)
            tk.Label(legend, text=ISSUE_LABELS[key], font=P.FONT_SMALL,
                     bg=P.BG, fg=P.FG_MUTED).pack(side="left", padx=(0, 16))

        detail = tk.Frame(bottom, bg=P.SURFACE, highlightthickness=1,
                          highlightbackground=P.BORDER)
        detail.pack(side="right", fill="x", expand=True, padx=(20, 0))
        tk.Label(detail, text="Row details", font=P.FONT_BOLD, bg=P.SURFACE,
                 fg=P.FG_STRONG).pack(anchor="w", padx=10, pady=(5, 0))
        self._detail = tk.Text(detail, height=5, font=P.FONT_SMALL, relief="flat",
                               bg=P.SURFACE, fg=P.FG, wrap="word",
                               state="disabled", highlightthickness=0)
        self._detail.pack(fill="x", padx=10, pady=(0, 8))
        self._detail.tag_configure("issue", foreground=P.ISSUE_FG["orange"])
        self._detail.tag_configure("key", foreground=P.ACCENT_TEXT)

    # ---------------- callbacks ----------------
    def _set_severity(self, option):
        self._sev_var.set(option)
        self.grid_view.set_severity(option)

    def _clear_filters(self):
        self._search_var.set("")
        self._sev_var.set(FILTER_OPTIONS[0])
        self.grid_view.clear_all_filters()

    def _on_view_change(self, shown, total):
        self._count_lbl.config(text=f"{shown:,} of {total:,} rows shown")

    def _update_kpis(self):
        rows = self.model["rows"]
        counts = {"yellow": 0, "red": 0, "orange": 0}
        for issues in self.model.get("row_issues", []):
            for k in issues:
                counts[k] += 1
        self._kpi_vals[None].config(text=f"{len(rows):,}")
        for k, v in counts.items():
            self._kpi_vals[k].config(text=f"{v:,}")

    def _show_detail(self, model_idx):
        self._detail.config(state="normal")
        self._detail.delete("1.0", "end")

        if model_idx is None or model_idx >= len(self.model["rows"]):
            self._detail.insert("1.0", "Select a row to see its full contents.")
            self._detail.config(state="disabled")
            return

        rows = self.model["rows"]
        cols = self.model["columns"]
        groups = self.model["groups"]
        hl = self.model["highlights"]

        def group_of(ci):
            for label, s, e in groups:
                if s <= ci <= e and label:
                    return label + " · "
            return ""

        issues = sorted(self.model["row_issues"][model_idx])
        if issues:
            self._detail.insert(
                "end", "⚠ " + ", ".join(ISSUE_LABELS[i] for i in issues) + "\n",
                "issue")

        for ci, v in enumerate(rows[model_idx]):
            if not v.strip():
                continue
            flag = f"  [{ISSUE_LABELS[hl[(model_idx, ci)]]}]" if (model_idx, ci) in hl else ""
            self._detail.insert("end", f"{group_of(ci)}{cols[ci]}: ", "key")
            self._detail.insert("end", f"{v.replace(chr(13), ' ').replace(chr(10), ' | ')}{flag}\n")

        self._detail.config(state="disabled")

    # ---------------- actions ----------------
    def _open_excel(self):
        src = self.model.get("source")
        if not src or not os.path.isfile(src):
            messagebox.showerror("Error", "Excel file not found.")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(src)                                   # noqa
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", src], check=False)
            else:
                import subprocess
                subprocess.run(["xdg-open", src], check=False)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}")

    def _export_csv(self):
        if not self.model["rows"]:
            messagebox.showinfo("Export", "There is nothing to export yet.")
            return
        path = filedialog.asksaveasfilename(
            title="Export visible rows",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")])
        if not path:
            return
        try:
            self.grid_view.export_view_csv(path)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{e}")
            return
        messagebox.showinfo("Export", "Visible rows exported.")


# ============================================================
# Convenience: results in their own window
# ============================================================
def open_results_window(parent, xlsx_path, title="Log Analysis Results"):
    win = tk.Toplevel(parent)
    win.title(title)
    win.configure(bg=P.BG)
    win.geometry("1360x780")
    win.minsize(940, 560)
    page = ResultsPage(win, xlsx_path=xlsx_path)
    page.pack(fill="both", expand=True)
    win.lift()
    return win


# ============================================================
# Standalone test:  python -m ui.results_page <output.xlsx>
# ============================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m ui.results_page <analysis_output.xlsx>")
        sys.exit(1)
    root = tk.Tk()
    root.title("Log Analysis Results")
    root.configure(bg=P.BG)
    root.geometry("1360x780")
    try:
        from ui.theme import apply_dark_theme
    except ImportError:
        from theme import apply_dark_theme
    apply_dark_theme(root)
    ResultsPage(root, xlsx_path=sys.argv[1]).pack(fill="both", expand=True)
    root.mainloop()
