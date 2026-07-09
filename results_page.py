# results_page.py
# ============================================================
# In-app results viewer for the log analyzer.
#
# Instead of (or in addition to) opening the generated Excel
# file, show the parsed data + flagged issues in a professional,
# fully custom page inside the app.
#
# It reads the .xlsx that analyze_logs() already produces, so
# NO changes to the analysis logic are needed:
#   - row 1  -> merged group headers (IXL Log, Wireshark Log, ...)
#   - row 2  -> column headers
#   - row 3+ -> data
#   - cell fills FFFF00 / FF0000 / FFA500 -> issue highlights
#
# Usage (from your app, after analyze_logs saves the file):
#
#   from results_page import open_results_window
#   open_results_window(root, output_file_path)
#
# or embed it as a page/frame like your other *_page.py modules:
#
#   from results_page import ResultsPage
#   page = ResultsPage(container, xlsx_path=output_file_path)
#   page.pack(fill="both", expand=True)
#
# Everything visual is driven by the THEME dict below, so the
# look can be tweaked freely (or overridden via theme= kwarg).
# ============================================================

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import load_workbook

# ------------------------------------------------------------
# Theme - edit freely (or pass a partial override dict)
# ------------------------------------------------------------
THEME = {
    "bg":             "#eef1f5",   # page background
    "surface":        "#ffffff",   # cards / table background
    "border":         "#d7dce3",

    "header_bg":      "#1f2937",   # column header
    "header_fg":      "#f9fafb",
    "group_bg":       "#111827",   # merged group header
    "group_fg":       "#e5e7eb",

    "row_even":       "#ffffff",
    "row_odd":        "#f6f8fa",
    "grid_line":      "#e5e7eb",
    "cell_fg":        "#1f2937",
    "muted_fg":       "#6b7280",

    "select_bg":      "#dbeafe",   # clicked row
    "accent":         "#2563eb",

    # issue colors (soft UI versions of the Excel fills)
    "issue_yellow":   "#fde68a",   # out-of-order sequence
    "issue_red":      "#fecaca",   # missing message
    "issue_orange":   "#fed7aa",   # missing packetswitch data

    "font":           ("Segoe UI", 9),
    "font_bold":      ("Segoe UI", 9, "bold"),
    "font_title":     ("Segoe UI", 14, "bold"),
    "font_kpi":       ("Segoe UI", 16, "bold"),
    "font_small":     ("Segoe UI", 8),

    "row_height":     24,
    "header_height":  26,
}

ISSUE_LABELS = {
    "yellow": "Out of Order Sequence",
    "red":    "Missing Message",
    "orange": "Missing Packetswitch Data",
}

FILTER_OPTIONS = [
    "All rows",
    "Issues only",
    "Out of Order Sequence",
    "Missing Message",
    "Missing Packetswitch Data",
]


# ------------------------------------------------------------
# Workbook -> model
# ------------------------------------------------------------
def _fill_to_issue(cell):
    """Map an openpyxl cell fill to an issue key ('yellow'/'red'/'orange')."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type != "solid":
            return None
        rgb = fill.start_color.rgb
        if not isinstance(rgb, str):
            return None
        rgb = rgb[-6:].upper()
        if rgb == "FFFF00":
            return "yellow"
        if rgb == "FF0000":
            return "red"
        if rgb == "FFA500":
            return "orange"
    except Exception:
        pass
    return None


def load_results_from_workbook(xlsx_path):
    """
    Parse the analyzer's output workbook into a plain model dict:
      columns    : [str, ...]                      (row 2)
      groups     : [(label, start_idx, end_idx)]   (row 1, 0-based, inclusive)
      rows       : [[str, ...], ...]               (row 3+)
      highlights : {(row_idx, col_idx): issue_key}
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    ncols = ws.max_column

    # --- group labels (row 1), expanded from merged ranges
    group_label = [""] * (ncols + 1)  # 1-based
    for rng in ws.merged_cells.ranges:
        if rng.min_row == 1 and rng.max_row == 1:
            val = ws.cell(row=1, column=rng.min_col).value
            for c in range(rng.min_col, rng.max_col + 1):
                group_label[c] = str(val) if val is not None else ""
    for c in range(1, ncols + 1):  # non-merged cells in row 1
        v = ws.cell(row=1, column=c).value
        if v is not None and not group_label[c]:
            group_label[c] = str(v)

    groups = []
    c = 1
    while c <= ncols:
        label = group_label[c]
        start = c
        while c + 1 <= ncols and group_label[c + 1] == label and label != "":
            c += 1
        groups.append((label, start - 1, c - 1))  # 0-based inclusive
        c += 1

    # --- column headers (row 2)
    columns = []
    for c in range(1, ncols + 1):
        v = ws.cell(row=2, column=c).value
        columns.append(str(v) if v is not None else "")

    # --- data rows + highlights
    rows, highlights = [], {}
    r_out = 0
    for r in range(3, ws.max_row + 1):
        vals, any_val = [], False
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            v = "" if cell.value is None else str(cell.value)
            if v.strip():
                any_val = True
            vals.append(v)
            issue = _fill_to_issue(cell)
            if issue:
                highlights[(r_out, c - 1)] = issue
        if not any_val:
            continue
        rows.append(vals)
        r_out += 1

    return {
        "columns": columns,
        "groups": groups,
        "rows": rows,
        "highlights": highlights,
        "source": os.path.abspath(xlsx_path),
    }


# ------------------------------------------------------------
# The page
# ------------------------------------------------------------
class ResultsPage(tk.Frame):
    def __init__(self, master, xlsx_path=None, model=None, theme=None, **kw):
        self.T = dict(THEME)
        if theme:
            self.T.update(theme)
        super().__init__(master, bg=self.T["bg"], **kw)

        if model is None:
            if xlsx_path is None:
                raise ValueError("ResultsPage needs xlsx_path or model")
            model = load_results_from_workbook(xlsx_path)
        self.model = model

        self._search_text = ""
        self._severity = FILTER_OPTIONS[0]
        self._sort_col = None       # column index
        self._sort_desc = False
        self._selected_view_row = None
        self._view = list(range(len(model["rows"])))  # filtered/sorted -> model idx

        self._col_widths = self._compute_col_widths()

        self._build_ui()
        self._apply_filters()

    # --------------- layout helpers ---------------
    def _compute_col_widths(self):
        f_px = 7  # approx px per char for 9pt UI font
        widths = []
        rows = self.model["rows"]
        sample = rows[:400]
        for ci, col in enumerate(self.model["columns"]):
            w = len(col)
            for row in sample:
                v = row[ci]
                if v:
                    first = v.splitlines()[0]
                    w = max(w, min(len(first), 42))
            widths.append(max(70, min(320, w * f_px + 18)))
        return widths

    # --------------- UI construction ---------------
    def _build_ui(self):
        T = self.T

        # ---- title bar
        title_fr = tk.Frame(self, bg=T["bg"])
        title_fr.pack(fill="x", padx=16, pady=(14, 6))
        tk.Label(title_fr, text="Analysis Results", font=T["font_title"],
                 bg=T["bg"], fg=T["cell_fg"]).pack(side="left")
        src = self.model.get("source", "")
        if src:
            tk.Label(title_fr, text=os.path.basename(src), font=T["font_small"],
                     bg=T["bg"], fg=T["muted_fg"]).pack(side="left", padx=(10, 0), pady=(6, 0))
        btn = tk.Button(title_fr, text="Open Excel file", font=T["font"],
                        relief="flat", bg=T["accent"], fg="#ffffff",
                        activebackground="#1d4ed8", activeforeground="#ffffff",
                        cursor="hand2", command=self._open_excel)
        btn.pack(side="right")

        # ---- KPI cards
        self._kpi_vals = {}
        kpi_fr = tk.Frame(self, bg=T["bg"])
        kpi_fr.pack(fill="x", padx=16, pady=(0, 8))
        cards = [
            ("Total Rows", None, T["cell_fg"], FILTER_OPTIONS[0]),
            (ISSUE_LABELS["yellow"], "yellow", "#92400e", ISSUE_LABELS["yellow"]),
            (ISSUE_LABELS["red"], "red", "#991b1b", ISSUE_LABELS["red"]),
            (ISSUE_LABELS["orange"], "orange", "#9a3412", ISSUE_LABELS["orange"]),
        ]
        for label, key, fg, filt in cards:
            card = tk.Frame(kpi_fr, bg=T["surface"], highlightthickness=1,
                            highlightbackground=T["border"], cursor="hand2")
            card.pack(side="left", padx=(0, 10), ipadx=14, ipady=6)
            accent = {"yellow": T["issue_yellow"], "red": T["issue_red"],
                      "orange": T["issue_orange"]}.get(key, T["accent"])
            tk.Frame(card, bg=accent, width=4).pack(side="left", fill="y")
            inner = tk.Frame(card, bg=T["surface"])
            inner.pack(side="left", padx=(10, 4))
            val = tk.Label(inner, text="0", font=T["font_kpi"], bg=T["surface"], fg=fg)
            val.pack(anchor="w")
            tk.Label(inner, text=label, font=T["font_small"], bg=T["surface"],
                     fg=T["muted_fg"]).pack(anchor="w")
            self._kpi_vals[key] = val
            for w in (card, inner, val):
                w.bind("<Button-1>", lambda e, f=filt: self._set_filter(f))

        # ---- filter bar
        bar = tk.Frame(self, bg=T["bg"])
        bar.pack(fill="x", padx=16, pady=(0, 8))

        tk.Label(bar, text="Search:", font=T["font"], bg=T["bg"],
                 fg=T["cell_fg"]).pack(side="left")
        self._search_var = tk.StringVar()
        ent = tk.Entry(bar, textvariable=self._search_var, font=T["font"], width=32,
                       relief="flat", highlightthickness=1,
                       highlightbackground=T["border"], highlightcolor=T["accent"])
        ent.pack(side="left", padx=(6, 16), ipady=3)
        self._search_var.trace_add("write", lambda *_: self._on_search())

        tk.Label(bar, text="Show:", font=T["font"], bg=T["bg"],
                 fg=T["cell_fg"]).pack(side="left")
        self._sev_var = tk.StringVar(value=FILTER_OPTIONS[0])
        combo = ttk.Combobox(bar, textvariable=self._sev_var, values=FILTER_OPTIONS,
                             state="readonly", width=26, font=T["font"])
        combo.pack(side="left", padx=(6, 16))
        combo.bind("<<ComboboxSelected>>", lambda e: self._set_filter(self._sev_var.get()))

        tk.Button(bar, text="Clear", font=T["font"], relief="flat",
                  bg=T["border"], cursor="hand2",
                  command=self._clear_filters).pack(side="left")

        self._count_lbl = tk.Label(bar, text="", font=T["font"], bg=T["bg"],
                                   fg=T["muted_fg"])
        self._count_lbl.pack(side="right")

        # ---- table
        table_wrap = tk.Frame(self, bg=T["border"], highlightthickness=1,
                              highlightbackground=T["border"])
        table_wrap.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        total_w = sum(self._col_widths)
        hdr_h = T["header_height"] * 2

        self._hdr = tk.Canvas(table_wrap, height=hdr_h, bg=T["header_bg"],
                              highlightthickness=0)
        self._hdr.grid(row=0, column=0, sticky="ew")
        self._body = tk.Canvas(table_wrap, bg=T["surface"], highlightthickness=0)
        self._body.grid(row=1, column=0, sticky="nsew")

        self._vsb = ttk.Scrollbar(table_wrap, orient="vertical",
                                  command=self._on_vscroll)
        self._vsb.grid(row=0, column=1, rowspan=2, sticky="ns")
        self._hsb = ttk.Scrollbar(table_wrap, orient="horizontal",
                                  command=self._on_hscroll)
        self._hsb.grid(row=2, column=0, sticky="ew")

        table_wrap.rowconfigure(1, weight=1)
        table_wrap.columnconfigure(0, weight=1)

        self._xoff = 0          # horizontal scroll offset (px)
        self._yrow = 0.0        # first visible row (float, in view coords)
        self._total_w = total_w

        self._body.bind("<Configure>", lambda e: self._redraw())
        self._body.bind("<MouseWheel>", self._on_wheel)           # Windows
        self._body.bind("<Button-4>", lambda e: self._wheel_lines(-3))  # X11
        self._body.bind("<Button-5>", lambda e: self._wheel_lines(3))
        self._body.bind("<Button-1>", self._on_body_click)
        self._hdr.bind("<Button-1>", self._on_header_click)

        # ---- legend + detail
        bottom = tk.Frame(self, bg=T["bg"])
        bottom.pack(fill="x", padx=16, pady=(0, 12))

        legend = tk.Frame(bottom, bg=T["bg"])
        legend.pack(side="left", anchor="n", pady=2)
        for key in ("yellow", "red", "orange"):
            sw = tk.Frame(legend, bg=self.T["issue_" + key], width=14, height=14,
                          highlightthickness=1, highlightbackground=T["border"])
            sw.pack(side="left", padx=(0, 4))
            sw.pack_propagate(False)
            tk.Label(legend, text=ISSUE_LABELS[key], font=T["font_small"], bg=T["bg"],
                     fg=T["muted_fg"]).pack(side="left", padx=(0, 14))

        detail_fr = tk.Frame(bottom, bg=T["surface"], highlightthickness=1,
                             highlightbackground=T["border"])
        detail_fr.pack(side="right", fill="x", expand=True, padx=(16, 0))
        tk.Label(detail_fr, text="Row details", font=T["font_bold"],
                 bg=T["surface"], fg=T["cell_fg"]).pack(anchor="w", padx=8, pady=(4, 0))
        self._detail = tk.Text(detail_fr, height=4, font=T["font_small"], relief="flat",
                               bg=T["surface"], fg=T["cell_fg"], wrap="word",
                               state="disabled")
        self._detail.pack(fill="x", padx=8, pady=(0, 6))

    # --------------- filtering / sorting ---------------
    def _row_issues(self, model_idx):
        ncols = len(self.model["columns"])
        hl = self.model["highlights"]
        return {hl[(model_idx, c)] for c in range(ncols) if (model_idx, c) in hl}

    def _set_filter(self, option):
        self._severity = option
        self._sev_var.set(option)
        self._apply_filters()

    def _clear_filters(self):
        self._search_var.set("")
        self._set_filter(FILTER_OPTIONS[0])

    def _on_search(self):
        self._search_text = self._search_var.get().strip().lower()
        self._apply_filters()

    def _apply_filters(self):
        rows = self.model["rows"]
        sev = self._severity
        needle = self._search_text

        idxs = []
        for i in range(len(rows)):
            if sev != "All rows":
                issues = self._row_issues(i)
                if sev == "Issues only":
                    if not issues:
                        continue
                else:
                    wanted = {v: k for k, v in ISSUE_LABELS.items()}.get(sev)
                    if wanted not in issues:
                        continue
            if needle:
                if not any(needle in v.lower() for v in rows[i] if v):
                    continue
            idxs.append(i)

        if self._sort_col is not None:
            col = self._sort_col

            def key(i):
                v = rows[i][col]
                try:
                    return (0, float(v))
                except (TypeError, ValueError):
                    return (1, v.lower())
            idxs.sort(key=key, reverse=self._sort_desc)

        self._view = idxs
        self._yrow = 0.0
        self._selected_view_row = None
        self._update_kpis()
        self._count_lbl.config(
            text=f"{len(idxs):,} of {len(rows):,} rows shown")
        self._redraw()

    def _update_kpis(self):
        rows = self.model["rows"]
        counts = {"yellow": 0, "red": 0, "orange": 0}
        for i in range(len(rows)):
            for issue in self._row_issues(i):
                counts[issue] += 1
        self._kpi_vals[None].config(text=f"{len(rows):,}")
        for k in counts:
            self._kpi_vals[k].config(text=f"{counts[k]:,}")

    # --------------- scrolling ---------------
    def _visible_rows(self):
        h = max(self._body.winfo_height(), 1)
        return max(int(h // self.T["row_height"]), 1)

    def _max_first_row(self):
        return max(len(self._view) - self._visible_rows(), 0)

    def _on_vscroll(self, *args):
        if args[0] == "moveto":
            frac = float(args[1])
            self._yrow = frac * self._max_first_row()
        elif args[0] == "scroll":
            step = int(args[1])
            if args[2] == "pages":
                step *= self._visible_rows()
            self._yrow += step
        self._yrow = min(max(self._yrow, 0), self._max_first_row())
        self._redraw()

    def _wheel_lines(self, n):
        self._yrow = min(max(self._yrow + n, 0), self._max_first_row())
        self._redraw()

    def _on_wheel(self, event):
        self._wheel_lines(-3 if event.delta > 0 else 3)

    def _on_hscroll(self, *args):
        w = max(self._body.winfo_width(), 1)
        max_off = max(self._total_w - w, 0)
        if args[0] == "moveto":
            self._xoff = int(float(args[1]) * max_off)
        elif args[0] == "scroll":
            step = int(args[1]) * (w if args[2] == "pages" else 40)
            self._xoff += step
        self._xoff = min(max(self._xoff, 0), max_off)
        self._redraw()

    def _update_scrollbars(self):
        n = max(len(self._view), 1)
        vis = self._visible_rows()
        first = self._yrow
        self._vsb.set(first / n, min((first + vis) / n, 1.0))

        w = max(self._body.winfo_width(), 1)
        if self._total_w <= w:
            self._hsb.set(0, 1)
        else:
            self._hsb.set(self._xoff / self._total_w,
                          (self._xoff + w) / self._total_w)

    # --------------- drawing ---------------
    def _redraw(self):
        self._draw_header()
        self._draw_body()
        self._update_scrollbars()

    def _draw_header(self):
        T = self.T
        c = self._hdr
        c.delete("all")
        hh = T["header_height"]
        x0 = -self._xoff

        # group row
        for label, s, e in self.model["groups"]:
            gx0 = x0 + sum(self._col_widths[:s])
            gx1 = gx0 + sum(self._col_widths[s:e + 1])
            c.create_rectangle(gx0, 0, gx1, hh, fill=T["group_bg"],
                               outline=T["header_bg"])
            if label:
                c.create_text((gx0 + gx1) / 2, hh / 2, text=label,
                              font=T["font_bold"], fill=T["group_fg"])

        # column row
        x = x0
        self._hdr_ranges = []
        for ci, (col, w) in enumerate(zip(self.model["columns"], self._col_widths)):
            c.create_rectangle(x, hh, x + w, hh * 2, fill=T["header_bg"],
                               outline=T["group_bg"])
            label = col
            if ci == self._sort_col:
                label += "  ▼" if self._sort_desc else "  ▲"
            c.create_text(x + w / 2, hh * 1.5, text=label, font=T["font_bold"],
                          fill=T["header_fg"], width=w - 10)
            self._hdr_ranges.append((x + self._xoff, x + self._xoff + w))
            x += w

    def _draw_body(self):
        T = self.T
        c = self._body
        c.delete("all")
        rh = T["row_height"]
        w_vis = c.winfo_width()
        vis = self._visible_rows() + 1

        first = int(self._yrow)
        rows = self.model["rows"]
        hl = self.model["highlights"]

        y = 0
        for vr in range(first, min(first + vis, len(self._view))):
            mi = self._view[vr]
            base = T["row_odd"] if (vr % 2) else T["row_even"]
            if vr == self._selected_view_row:
                base = T["select_bg"]
            c.create_rectangle(-self._xoff, y, -self._xoff + self._total_w, y + rh,
                               fill=base, outline="")
            x = -self._xoff
            for ci, w in enumerate(self._col_widths):
                if x + w >= 0 and x <= w_vis:  # draw only visible cells
                    issue = hl.get((mi, ci))
                    if issue:
                        c.create_rectangle(x + 1, y + 1, x + w - 1, y + rh - 1,
                                           fill=T["issue_" + issue], outline="")
                    v = rows[mi][ci]
                    if v:
                        text = v.splitlines()[0]
                        if len(text) > 44:
                            text = text[:43] + "…"
                        c.create_text(x + 6, y + rh / 2, text=text, anchor="w",
                                      font=T["font"], fill=T["cell_fg"])
                x += w
            c.create_line(-self._xoff, y + rh, -self._xoff + self._total_w, y + rh,
                          fill=T["grid_line"])
            y += rh

        # vertical grid lines
        x = -self._xoff
        for w in self._col_widths:
            x += w
            c.create_line(x, 0, x, y, fill=T["grid_line"])

    # --------------- interactions ---------------
    def _on_header_click(self, event):
        hh = self.T["header_height"]
        if event.y < hh:      # click on group row -> ignore
            return
        x = event.x + self._xoff
        acc = 0
        for ci, w in enumerate(self._col_widths):
            if acc <= x < acc + w:
                if self._sort_col == ci:
                    if self._sort_desc:
                        self._sort_col = None   # third click clears sort
                        self._sort_desc = False
                    else:
                        self._sort_desc = True
                else:
                    self._sort_col, self._sort_desc = ci, False
                self._apply_filters()
                return
            acc += w

    def _on_body_click(self, event):
        vr = int(self._yrow) + event.y // self.T["row_height"]
        if vr >= len(self._view):
            return
        self._selected_view_row = vr
        mi = self._view[vr]
        self._show_detail(mi)
        self._redraw()

    def _show_detail(self, mi):
        rows = self.model["rows"]
        cols = self.model["columns"]
        groups = self.model["groups"]
        hl = self.model["highlights"]

        def group_of(ci):
            for label, s, e in groups:
                if s <= ci <= e and label:
                    return label + " · "
            return ""

        parts = []
        issues = sorted(self._row_issues(mi))
        if issues:
            parts.append("⚠ " + ", ".join(ISSUE_LABELS[i] for i in issues))
        for ci, v in enumerate(rows[mi]):
            if not v.strip():
                continue
            flag = ""
            if (mi, ci) in hl:
                flag = f"  [{ISSUE_LABELS[hl[(mi, ci)]]}]"
            parts.append(f"{group_of(ci)}{cols[ci]}: {v}{flag}")
        text = "\n".join(parts)

        self._detail.config(state="normal")
        self._detail.delete("1.0", "end")
        self._detail.insert("1.0", text)
        self._detail.config(state="disabled")

    def _open_excel(self):
        src = self.model.get("source")
        if not src or not os.path.isfile(src):
            messagebox.showerror("Error", "Excel file not found.")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(src)                      # noqa
            elif sys.platform == "darwin":
                os.system(f'open "{src}"')
            else:
                os.system(f'xdg-open "{src}"')
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {e}")


# ------------------------------------------------------------
# Convenience: results in their own window
# ------------------------------------------------------------
def open_results_window(parent, xlsx_path, theme=None, title="Log Analysis Results"):
    win = tk.Toplevel(parent)
    win.title(title)
    win.geometry("1280x760")
    win.minsize(900, 520)
    page = ResultsPage(win, xlsx_path=xlsx_path, theme=theme)
    page.pack(fill="both", expand=True)
    win.lift()
    return win


# ------------------------------------------------------------
# Standalone test:  python results_page.py <output.xlsx>
# ------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python results_page.py <analysis_output.xlsx>")
        sys.exit(1)
    root = tk.Tk()
    root.title("Log Analysis Results")
    root.geometry("1280x760")
    ResultsPage(root, xlsx_path=sys.argv[1]).pack(fill="both", expand=True)
    root.mainloop()
